#!/usr/bin/env python3
"""Creates the initial tracking.xlsx for Experiment 1."""
import os, sys
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "../..")))

from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
from data import MBTI_PERSONAS

TYPES  = list(MBTI_PERSONAS.keys())
OUT    = os.path.join(os.path.dirname(__file__), "tracking.xlsx")

# ── colours ──────────────────────────────────────────────────────────
C_HEADER  = "1F3864"   # dark navy
C_PENDING = "2D2D2D"   # dark grey
C_DONE    = "1A4D1A"   # dark green
C_ERROR   = "6B1A1A"   # dark red
C_RUNNING = "4D3A00"   # dark amber
C_ALT     = "1A1A2E"   # alternate row

THIN = Border(
    left=Side(style="thin", color="333333"),
    right=Side(style="thin", color="333333"),
    top=Side(style="thin", color="333333"),
    bottom=Side(style="thin", color="333333"),
)

def hdr_fill(hex_col): return PatternFill("solid", fgColor=hex_col)
def hdr_font(bold=True): return Font(name="Calibri", bold=bold, color="FFFFFF", size=11)
def cell_font(): return Font(name="Calibri", color="CCCCCC", size=10)

wb = Workbook()

# ════════════════════════════════════════════════════════════════════
# Sheet 1 — flat experiment list
# ════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Experiments"
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A2"

headers = ["ID", "Persona 1", "Persona 2", "Status",
           "Scenario", "Log File", "Start Time", "End Time", "Turns",
           "Tokens In", "Tokens Out", "Cost ($)"]
col_widths = [6, 10, 10, 12, 52, 38, 20, 20, 8, 12, 12, 10]

for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
    c = ws.cell(row=1, column=ci, value=h)
    c.font      = hdr_font()
    c.fill      = hdr_fill(C_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = THIN
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.row_dimensions[1].height = 22

pairs = [(p1, p2) for p1 in TYPES for p2 in TYPES]
for idx, (p1, p2) in enumerate(pairs, 1):
    row = idx + 1
    fill = hdr_fill(C_ALT if idx % 2 == 0 else C_PENDING)
    vals = [idx, p1, p2, "pending", "", "", "", "", ""]
    for ci, val in enumerate(vals, 1):
        c = ws.cell(row=row, column=ci, value=val)
        c.font      = cell_font()
        c.fill      = fill
        c.alignment = Alignment(horizontal="center" if ci in (1,9) else "left",
                                vertical="center")
        c.border    = THIN
    ws.row_dimensions[row].height = 16

# ════════════════════════════════════════════════════════════════════
# Sheet 2 — 16×16 grid
# ════════════════════════════════════════════════════════════════════
wg = wb.create_sheet("Grid")
wg.sheet_view.showGridLines = False
wg.freeze_panes = "B2"

# top-left corner label
corner = wg.cell(row=1, column=1, value="P1 \\ P2")
corner.font      = hdr_font()
corner.fill      = hdr_fill(C_HEADER)
corner.alignment = Alignment(horizontal="center", vertical="center")
corner.border    = THIN

# column headers (Persona 2)
for ci, t in enumerate(TYPES, 2):
    c = wg.cell(row=1, column=ci, value=t)
    c.font      = hdr_font()
    c.fill      = hdr_fill(C_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = THIN
    wg.column_dimensions[get_column_letter(ci)].width = 9
wg.column_dimensions["A"].width = 9
wg.row_dimensions[1].height = 22

# row headers (Persona 1) + cells
for ri, p1 in enumerate(TYPES, 2):
    rh = wg.cell(row=ri, column=1, value=p1)
    rh.font      = hdr_font()
    rh.fill      = hdr_fill(C_HEADER)
    rh.alignment = Alignment(horizontal="center", vertical="center")
    rh.border    = THIN
    wg.row_dimensions[ri].height = 18

    for ci, p2 in enumerate(TYPES, 2):
        c = wg.cell(row=ri, column=ci, value="pending")
        c.font      = Font(name="Calibri", color="888888", size=9)
        c.fill      = hdr_fill(C_PENDING)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = THIN

# ════════════════════════════════════════════════════════════════════
# Sheet 3 — batch plan
# ════════════════════════════════════════════════════════════════════
wb3 = wb.create_sheet("Batch Plan")
wb3.sheet_view.showGridLines = False

bh = ["Batch", "Battles", "Persona Pairs"]
bw = [8, 10, 80]
for ci, (h, w) in enumerate(zip(bh, bw), 1):
    c = wb3.cell(row=1, column=ci, value=h)
    c.font      = hdr_font()
    c.fill      = hdr_fill(C_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = THIN
    wb3.column_dimensions[get_column_letter(ci)].width = w
wb3.row_dimensions[1].height = 22

BATCH_SIZE = 8
for b_idx, start in enumerate(range(0, len(pairs), BATCH_SIZE), 1):
    batch = pairs[start:start+BATCH_SIZE]
    label = ", ".join(f"{a}↔{b}" for a, b in batch)
    row   = b_idx + 1
    fill  = hdr_fill(C_ALT if b_idx % 2 == 0 else C_PENDING)
    for ci, val in enumerate([b_idx, len(batch), label], 1):
        c = wb3.cell(row=row, column=ci, value=val)
        c.font      = cell_font()
        c.fill      = fill
        c.alignment = Alignment(horizontal="center" if ci < 3 else "left", vertical="center")
        c.border    = THIN
    wb3.row_dimensions[row].height = 16

wb.save(OUT)
print(f"Created: {OUT}")
print(f"  Experiments sheet : {len(pairs)} rows")
print(f"  Grid sheet        : 16×16")
print(f"  Batch Plan sheet  : {b_idx} batches of up to {BATCH_SIZE}")
