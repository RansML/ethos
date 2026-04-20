#!/usr/bin/env python3
"""
Experiment 1 — 16 vs 16 MBTI Battle
All 256 persona pairs (16×16), run 8 at a time in parallel.
Results saved to experiments/experiment_1/chats/ and tracked in tracking.xlsx.

Usage:
    python experiments/experiment_1/run_experiment.py
"""
import os
import sys
import random
import time
import traceback
from datetime import datetime
from concurrent.futures import ProcessPoolExecutor, as_completed

import openai

# ── paths ─────────────────────────────────────────────────────────────
ROOT        = os.path.abspath(os.path.join(os.path.dirname(__file__), "../.."))
_data_dir = f"/data/{os.environ.get('USER', 'user')}/ethos/experiments/experiment_1/results"
CHATS_DIR = _data_dir if os.path.exists("/data") else os.path.join(os.path.dirname(__file__), "results")
TRACKING_XL = os.path.join(os.path.dirname(__file__), "tracking.xlsx")
ERROR_LOG   = os.path.join(os.path.dirname(__file__), "errors.log")
sys.path.insert(0, ROOT)

from dotenv import load_dotenv
load_dotenv(os.path.join(ROOT, ".env"))

from data import MBTI_PERSONAS, SCENARIO_SEEDS, build_system_prompt
from openai import OpenAI

TYPES       = list(MBTI_PERSONAS.keys())          # 16 types
MAX_TURNS   = 50
BATCH_SIZE  = 8
SCENARIO    = SCENARIO_SEEDS[0]                   # "a tense family dinner where something unexpected was just revealed"


# ── all 256 pairs ─────────────────────────────────────────────────────
def all_pairs() -> list[tuple[str, str]]:
    return [(p1, p2) for p1 in TYPES for p2 in TYPES]


# ── single battle ─────────────────────────────────────────────────────
def run_battle(pair: tuple[str, str]) -> dict:
    p1, p2 = pair
    os.makedirs(CHATS_DIR, exist_ok=True)

    scenario   = SCENARIO
    timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_name   = f"{p1}_vs_{p2}_{timestamp}.txt"
    log_path   = os.path.join(CHATS_DIR, log_name)

    sys1 = build_system_prompt(p1, MBTI_PERSONAS[p1], scenario) + \
           f"\n\nYou are talking with {p2}. Keep it short and casual."
    sys2 = build_system_prompt(p2, MBTI_PERSONAS[p2], scenario) + \
           f"\n\nYou are talking with {p1}. Keep it short and casual."

    client       = OpenAI()
    history      = []   # [{"speaker": "1"|"2", "content": str}]
    tokens_in    = 0
    tokens_out   = 0

    def msgs_for(perspective: str) -> list:
        out = []
        for e in history:
            out.append({"role": "assistant" if e["speaker"] == perspective else "user",
                        "content": e["content"]})
        return out

    def reply(perspective: str, max_tok: int = 256) -> str:
        nonlocal tokens_in, tokens_out
        sys_p = sys1 if perspective == "1" else sys2
        r = client.chat.completions.create(
            model="gpt-5.4-mini",
            max_completion_tokens=max_tok,
            messages=[{"role": "system", "content": sys_p}] + msgs_for(perspective),
        )
        tokens_in  += r.usage.prompt_tokens
        tokens_out += r.usage.completion_tokens
        return r.choices[0].message.content.strip()

    start = datetime.now()

    try:
        with open(log_path, "w") as f:
            f.write(f"Experiment 1 — {p1} vs {p2}\n")
            f.write(f"Scenario: {scenario}\n")
            f.write(f"Started: {start.strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("=" * 60 + "\n\n")

            seed = [{"role": "user", "content":
                     "Open the conversation naturally, in character, already in this situation. One or two casual sentences."}]
            r0 = client.chat.completions.create(
                model="gpt-5.4-mini", max_completion_tokens=150,
                messages=[{"role": "system", "content": sys1}] + seed,
            )
            tokens_in  += r0.usage.prompt_tokens
            tokens_out += r0.usage.completion_tokens
            opening = r0.choices[0].message.content.strip()

            history.append({"speaker": "1", "content": opening})
            f.write(f"[{p1}] {opening}\n")

            while len(history) < MAX_TURNS:
                last      = history[-1]["speaker"]
                next_sp   = "2" if last == "1" else "1"
                next_name = p2 if next_sp == "2" else p1
                text      = reply(next_sp)
                history.append({"speaker": next_sp, "content": text})
                f.write(f"[{next_name}] {text}\n")
                f.flush()

            end = datetime.now()
            f.write(f"\n[ended: {end.strftime('%Y-%m-%d %H:%M:%S')}]\n")

    except openai.APIError as e:
        raise RuntimeError(f"OpenAI API error [{type(e).__name__}]: {e}") from e
    except openai.RateLimitError as e:
        raise RuntimeError(f"Rate limit hit: {e}") from e
    except openai.AuthenticationError as e:
        raise RuntimeError(f"Auth error (check API key): {e}") from e
    except Exception as e:
        raise RuntimeError(f"Unexpected error: {type(e).__name__}: {e}\n{traceback.format_exc()}") from e

    return {
        "p1": p1, "p2": p2,
        "scenario": scenario,
        "turns": len(history),
        "log": log_name,
        "start": start.strftime("%Y-%m-%d %H:%M:%S"),
        "end": end.strftime("%Y-%m-%d %H:%M:%S"),
        "status": "done",
        "tokens_in": tokens_in,
        "tokens_out": tokens_out,
    }


# ── excel update ──────────────────────────────────────────────────────
def update_excel(results: list[dict]):
    from openpyxl import load_workbook
    wb = load_workbook(TRACKING_XL)

    # ── Sheet 1: flat list ─────────────────────────────────────────
    ws = wb["Experiments"]
    lookup = {(r["p1"], r["p2"]): r for r in results}

    for row in ws.iter_rows(min_row=2):
        p1_cell, p2_cell = row[1].value, row[2].value
        if (p1_cell, p2_cell) in lookup:
            r = lookup[(p1_cell, p2_cell)]
            row[3].value = r["status"]
            row[4].value = r["scenario"]
            row[5].value = r["log"]
            row[6].value = r["start"]
            row[7].value = r["end"]
            row[8].value = r["turns"]
            row[9].value  = r.get("tokens_in", 0)
            row[10].value = r.get("tokens_out", 0)
            row[11].value = round(
                (r.get("tokens_in", 0) / 1_000_000) * COST_PER_M_IN +
                (r.get("tokens_out", 0) / 1_000_000) * COST_PER_M_OUT, 6
            )

    # ── Sheet 2: grid ─────────────────────────────────────────────
    wg = wb["Grid"]
    for r in results:
        ri = TYPES.index(r["p1"]) + 2
        ci = TYPES.index(r["p2"]) + 2
        wg.cell(row=ri, column=ci).value = r["status"]

    wb.save(TRACKING_XL)


# gpt-5.4-mini pricing (per 1M tokens)
COST_PER_M_IN  = 0.150
COST_PER_M_OUT = 0.600


def update_config(exp_start: datetime, exp_end: datetime, results: list[dict]):
    total_in   = sum(r.get("tokens_in",  0) for r in results)
    total_out  = sum(r.get("tokens_out", 0) for r in results)
    total_tok  = total_in + total_out
    cost       = (total_in / 1_000_000) * COST_PER_M_IN + (total_out / 1_000_000) * COST_PER_M_OUT
    duration   = exp_end - exp_start
    h, rem     = divmod(int(duration.total_seconds()), 3600)
    m, s       = divmod(rem, 60)

    config_path = os.path.join(os.path.dirname(__file__), "config.md")
    with open(config_path, "r") as f:
        content = f.read()

    run_section = f"""
## Run Results
| Metric | Value |
|--------|-------|
| Experiment start | {exp_start.strftime("%Y-%m-%d %H:%M:%S")} |
| Experiment end   | {exp_end.strftime("%Y-%m-%d %H:%M:%S")} |
| Total duration   | {h}h {m}m {s}s |
| Battles completed | {len(results)} / 256 |
| Total turns | {sum(r.get("turns", 0) for r in results)} |
| Tokens in  | {total_in:,} |
| Tokens out | {total_out:,} |
| Total tokens | {total_tok:,} |
| Estimated cost | ${cost:.4f} |
"""

    marker = "\n## Run Results"
    if marker in content:
        content = content[:content.index(marker)]
    content = content.rstrip() + "\n" + run_section

    with open(config_path, "w") as f:
        f.write(content)

    print(f"\n  Duration : {h}h {m}m {s}s")
    print(f"  Tokens   : {total_tok:,} (in: {total_in:,} / out: {total_out:,})")
    print(f"  Cost     : ${cost:.4f}")


# ── main ─────────────────────────────────────────────────────────────
def main():
    pairs     = all_pairs()
    total     = len(pairs)
    exp_start = datetime.now()
    print(f"\nExperiment 1 — {total} battles ({BATCH_SIZE} parallel)\n")

    all_results = []
    batches     = [pairs[i:i+BATCH_SIZE] for i in range(0, total, BATCH_SIZE)]

    for b_idx, batch in enumerate(batches):
        print(f"Batch {b_idx+1}/{len(batches)} — running {len(batch)} battles...")
        batch_results = []

        failed = False
        with ProcessPoolExecutor(max_workers=BATCH_SIZE) as ex:
            futures = {ex.submit(run_battle, pair): pair for pair in batch}
            for fut in as_completed(futures):
                pair = futures[fut]
                p1, p2 = pair
                try:
                    result = fut.result()
                    batch_results.append(result)
                    print(f"  ✓  {p1} vs {p2} — {result['turns']} turns")
                except Exception as e:
                    err_msg = (
                        f"\n[ERROR] Batch {b_idx+1} — {p1} vs {p2}\n"
                        f"Time   : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
                        f"Error  : {e}\n"
                        f"{'─'*60}\n"
                    )
                    print(f"  ✗  {p1} vs {p2} — {e}")
                    print("\n[STOPPING] Error encountered. Saving progress and exiting.")
                    with open(ERROR_LOG, "a") as ef:
                        ef.write(err_msg)
                    # cancel remaining futures
                    for f2 in futures:
                        f2.cancel()
                    failed = True
                    break

        all_results.extend(batch_results)
        update_excel(batch_results)
        print(f"  → tracking.xlsx updated\n")

        if failed:
            print(f"Error log : {ERROR_LOG}\n")
            sys.exit(1)

    exp_end = datetime.now()
    update_config(exp_start, exp_end, all_results)

    print(f"\nDone. {len(all_results)} battles completed.")
    print(f"Tracking : {TRACKING_XL}")
    print(f"Logs     : {CHATS_DIR}\n")


if __name__ == "__main__":
    main()
